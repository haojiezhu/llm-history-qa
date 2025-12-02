import os, random
from ollama import Client
from openpyxl import Workbook

# LLM name for tracking inference results
llm_name = "gpt_oss_20b"

# Use default TCP/IP port (port 11434) for Ollama server
#os.system("./ollama/bin/ollama pull gpt-oss:20b")
#client = Client()

# Use custom TCP/IP port (port 11435 in this example) for Ollama server
os.system("OLLAMA_HOST=127.0.0.1:11435 ./ollama/bin/ollama pull gpt-oss:20b")
client = Client(host="127.0.0.1:11435")

qa_dataset = []

file_path = "./history_qa_csv.csv"
with open(file_path, 'r') as file:
	lines = file.readlines()
	for line in lines:
		data_fields = line.strip().split("|")
		if data_fields[0] != "Template_ID":
			template_id = data_fields[0]
			qa_template = data_fields[1]
			qa_id = data_fields[2]
			num_placeholders = int(data_fields[3])
			num_choices = int(data_fields[4])
			ground_truth = data_fields[5]
			# Parse all MCQ choices
			all_choices = data_fields[6].strip().split(";")
			for i in range(len(all_choices)):
				all_choices[i] = all_choices[i].strip()
			# Randomizing choices and change ground_truth accordingly
			if num_choices == 4 and ground_truth in ["A", "B", "C", "D"]:
				# Example: "C" is ground truth -> ground_truth_index = 2
				dict_letter_index = {"A":0, "B":1, "C":2, "D":3}
				ground_truth_index = dict_letter_index[ground_truth]
				indices = [0, 1, 2, 3]
				random.shuffle(indices)
				shuffled_choices = [all_choices[indices[0]], all_choices[indices[1]], all_choices[indices[2]], all_choices[indices[3]]]
				# Find new ground truth position after shuffling, e.g. [1, 3, 0, 2]
				ground_truth_index_new = indices.index(ground_truth_index)
				dict_index_letter = {0:"A", 1:"B", 2:"C", 3:"D"}
				# Update ground truth
				ground_truth = dict_index_letter[ground_truth_index_new]
				# Update all choices
				all_choices = shuffled_choices
			elif num_choices == 2 and ground_truth in ["A", "B"]:
				# Example: "B" is ground truth -> ground_truth_index = 1
				dict_letter_index = {"A":0, "B":1}
				ground_truth_index = dict_letter_index[ground_truth]
				indices = [0, 1]
				random.shuffle(indices)
				shuffled_choices = [all_choices[indices[0]], all_choices[indices[1]]]
				# Find new ground truth position after shuffling, e.g. [1, 0]
				ground_truth_index_new = indices.index(ground_truth_index)
				dict_index_letter = {0:"A", 1:"B"}
				# Update ground truth
				ground_truth = dict_index_letter[ground_truth_index_new]
				# Update all choices
				all_choices = shuffled_choices
			# Insert placeholders into template
			for i in range(num_placeholders):
				str_to_replace = "[p" + str(i+1) + "]"
				qa_template = qa_template.replace(str_to_replace, data_fields[i+7])
			#print("{0}, {1}, {2}".format(qa_id, template_id, qa_template))
			#print("{0}, {1}".format(ground_truth, all_choices))
			# Construct final LLM QA question
			if num_choices == 4:
				prefix = template_id + " | " + qa_id + " | You are an expert on historical events. Please select the best choice for the following multiple choice question. Don\'t show any reasoning process. Just give me the letter of your best choice.\n"
				suffix = "\nA. " + all_choices[0] + "\nB. " + all_choices[1] + "\nC. " + all_choices[2] + "\nD. " + all_choices[3] + " | " + ground_truth
				llm_qa_question = prefix + qa_template + suffix
			elif num_choices == 2 and ground_truth in ["TRUE", "FALSE"]:
				prefix = template_id + " | " + qa_id + " | You are an expert on historical events. Please answer the following true or false question. Don\'t show any reasoning process. Just give me TRUE or FALSE as your answer. Here is the question:\n"
				suffix = " | " + ground_truth
				llm_qa_question = prefix + qa_template + suffix
			elif num_choices == 2 and ground_truth in ["A", "B"]:
				prefix = template_id + " | " + qa_id + " | You are an expert on historical events. Please select the best choice for the following multiple choice question. Don\'t show any reasoning process. Just give me the letter of your best choice.\n"
				suffix = "\nA. " + all_choices[0] + "\nB. " + all_choices[1] + " | " + ground_truth
				llm_qa_question = prefix + qa_template + suffix
			else:
				print("Can't process question " + qa_id + " with ground truth: " + ground_truth)
				continue
			# Add question with proper LLM prompt as a row in dataset
			print(llm_qa_question)
			qa_dataset.append(llm_qa_question)

list_results = []

for qa_row in qa_dataset:
	qa_question = qa_row.strip().split("|")
	resp_fast = client.chat(model="gpt-oss:20b", messages=[{"role": "user", "content": qa_question[2].strip()}], think="low")
	list_results.append({"template_id": qa_question[0].strip(), "question_id": qa_question[1].strip(), "question_text": qa_question[2].strip(), "ground_truth": qa_question[3].strip(), "model_inference": resp_fast["message"]["content"]})

# Print out QA dataset
print("template_id\t" + "question_id\t" + "question_text")
for result in list_results:
	print("-------------------------------")
	print(result["template_id"] + "\t" + result["question_id"] + "\t" + result["question_text"])

print()

# Print out ground truth vs. inference
print("question_id\t" + "ground_truth\t" + "model_inference")

for result in list_results:
	# Model inference may return extra texts
	if len(result["model_inference"]) >=2 and result["model_inference"][:2] in ["A)", "B)", "C)", "D)"]:
		model_inference = result["model_inference"][0]
	else:
		model_inference = result["model_inference"]
	print(result["question_id"] + "\t" + result["ground_truth"] + "\t" + model_inference)

# Export inference results to an Excel spreadsheet

# Create a new workbook
wb = Workbook()

# Get the active worksheet
ws = wb.active
ws.title = "Results"  # Set the sheet title

# Write headers
ws['A1'] = "Template_ID"
ws['B1'] = "Question_ID"
ws['C1'] = "LLM_Name"
ws['D1'] = "Ground_Truth"
ws['E1'] = "Model_Inference"

for result in list_results:
	# Model inference may return extra texts
	if len(result["model_inference"]) >=2 and result["model_inference"][:2] in ["A)", "B)", "C)", "D)"]:
		model_inference = result["model_inference"][0]
	else:
		model_inference = result["model_inference"]

	# Append a new row
	data_row = [result["template_id"], result["question_id"], llm_name, result["ground_truth"], model_inference]
	ws.append(data_row)

# Save the workbook
excel_file_path = "./results_" + llm_name + ".xlsx"
wb.save(excel_file_path)
print("Data written to '{0}' successfully.".format(excel_file_path))

