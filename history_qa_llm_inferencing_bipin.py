import os
from ollama import Client
from openpyxl import load_workbook, Workbook

# LLM name for tracking inference results
llm_name = "gpt_oss_20b"

# Use default TCP/IP port (port 11434) for Ollama server
os.system("./ollama/bin/ollama pull gpt-oss:20b")
client = Client()

# Use custom TCP/IP port (port 11435 in this example) for Ollama server
#os.system("OLLAMA_HOST=127.0.0.1:11435 ./ollama/bin/ollama pull gpt-oss:20b")
#client = Client(host="127.0.0.1:11435")

qa_dataset = []

file_path = "./history_qa_bipin.xlsx"

# Read from Excel spreadsheet directly
workbook = load_workbook(filename=file_path)
sheet = workbook.active

for row in sheet.iter_rows(min_row=2, max_row=201, min_col=1, max_col=4):
	template_id = str(row[0].value)
	qa_id = str(row[1].value)
	qa_text = row[2].value
	ground_truth = str(row[3].value).upper()
	if ground_truth in ["TRUE", "FALSE"]:
		prefix = template_id + " | " + qa_id + " | You are an expert on historical events and tasked with verifying whether the following claim is historically accurate. Don\'t show any reasoning process. Just give me TRUE or FALSE as your answer. Here is the claim:\n"
		suffix = " | " + ground_truth
		llm_qa_question = prefix + qa_text + suffix
		#print(llm_qa_question)
	elif ground_truth in ["A", "B", "C", "D"]:
		prefix = template_id + " | " + qa_id + " | You are an expert on historical events. Please select the best choice for the following multiple choice question. Don\'t show any reasoning process. Just give me the letter of your best choice.\n"
		suffix = " | " + ground_truth
		llm_qa_question = prefix + qa_text + suffix
		#print(llm_qa_question)
	else:
		print("Can't process question " + qa_id + " with ground truth: " + ground_truth)
		continue
	# Add question with proper LLM prompt as a row in dataset
	qa_dataset.append(llm_qa_question)

list_results = []

# Running LLM inference on Ollama server
for qa_row in qa_dataset:
	qa_question = qa_row.strip().split("|")
	resp_fast = client.chat(model="gpt-oss:20b", messages=[{"role": "user", "content": qa_question[2].strip()}], think="low")
	list_results.append({"template_id": qa_question[0].strip(), "question_id": qa_question[1].strip(), "question_text": qa_question[2].strip(), "ground_truth": qa_question[3].strip(), "model_inference": resp_fast["message"]["content"]})

# Print out QA dataset
print("template_id\t" + "question_id\t" + "question_text")
for result in list_results:
	print("-------------------------------")
	print(result["template_id"] + "\t" + result["question_id"] + "\t" + result["question_text"])

print()

# Print out ground truth vs. inference
print("question_id\t" + "ground_truth\t" + "model_inference")

for result in list_results:
	# Model inference may return extra texts
	if len(result["model_inference"]) >=2 and result["model_inference"][:2] in ["A)", "B)", "C)", "D)"]:
		model_inference = result["model_inference"][0]
	else:
		model_inference = result["model_inference"]
	print(result["question_id"] + "\t" + result["ground_truth"] + "\t" + model_inference)

# Export inference results to an Excel spreadsheet

# Create a new workbook
wb = Workbook()

# Get the active worksheet
ws = wb.active
ws.title = "Results"  # Set the sheet title

# Write headers
ws['A1'] = "Template_ID"
ws['B1'] = "Question_ID"
ws['C1'] = "LLM_Name"
ws['D1'] = "Ground_Truth"
ws['E1'] = "Model_Inference"

for result in list_results:
	# Model inference may return extra texts
	if len(result["model_inference"]) >=2 and result["model_inference"][:2] in ["A)", "B)", "C)", "D)"]:
		model_inference = result["model_inference"][0]
	else:
		model_inference = result["model_inference"]

	# Append a new row
	data_row = [result["template_id"], result["question_id"], llm_name, result["ground_truth"], model_inference]
	ws.append(data_row)

# Save the workbook
excel_file_path = "./results_" + llm_name + "_bipin.xlsx"
wb.save(excel_file_path)
print("Data written to '{0}' successfully.".format(excel_file_path))
